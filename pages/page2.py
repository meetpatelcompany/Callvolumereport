import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
import os



def process_excel_file(file_path, sheet_name, date_column='Date'):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Delete the first row
    ws.delete_rows(1)
    
    # Identify the column index for the 'Date' column
    date_col_index = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == date_column:
            date_col_index = col_idx
            break

    if date_col_index is None:
        raise ValueError(f"Column '{date_column}' not found in the sheet.")

    # List to keep track of rows to delete
    rows_to_delete = set()
    
    # Iterate through rows to find empty cells in the 'Date' column
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=date_col_index, max_col=date_col_index, values_only=True), start=2):
        if row[0] is None:
            rows_to_delete.update(range(row_idx, min(row_idx + 6, ws.max_row + 1)))  # Add 5 more rows to the deletion set
    
    # Delete rows in reverse order to avoid affecting the indices of rows yet to be deleted
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    # Save the changes to the Excel file
    wb.save(file_path)

def filter_dataframe_with_openpyxl(file_path, sheet_name, conditions, exclude_value=None, exclude_column=None):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    
    # Exclude specific value if provided
    if exclude_value is not None and exclude_column is not None:
        df = df[df[exclude_column] != exclude_value]
    
    filtered_df = df[conditions]
    
    # Clear the worksheet and write the filtered data back
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(file_path)

# Function to generate dates between start and end date
def generate_dates(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)

# Function to find the first empty row in a column
def find_first_empty_row(ws, column):
    for row in range(1, ws.max_row + 1):
        if ws[f'{column}{row}'].value is None:
            return row
    return ws.max_row + 1

# Function to copy data from DataFrame to Excel
def copy_data_to_excel(df, ws, df_column, excel_column, font):
    first_empty_row = find_first_empty_row(ws, excel_column)
    for i, value in enumerate(df[df_column], start=first_empty_row):
        cell = ws[f'{excel_column}{i}']
        cell.value = value
        cell.font = font

def find_stop_row(sheet, col):
    for i, cell in enumerate(sheet[col], start=1):
        if cell.value is None or cell.value == "":
            return i - 1
    return len(sheet[col])
st.subheader("Instructions for Page 2",divider=True)

st.markdown("This process devide into 2 parts")
st.markdown("1. Choose Start and end Date for Previous week")
st.markdown("2. Press Execute button and wait till process complete")
# st.markdown("3. Press Execute button and wait till process complete")
# st.markdown("4. You can find modified files in :blue-background[\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet]")
start_date_str = st.date_input('Start Date', datetime(2024, 7, 15))
end_date_str = st.date_input('End Date', datetime(2024, 7, 21))

file_path = r'\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet\Call Volume Data - Base Data - Working - CUIC.xlsx'

if st.button('Execute'):
                # Convert start and end date to datetime objects
    start_date = pd.to_datetime(start_date_str, errors='coerce')
    end_date = pd.to_datetime(end_date_str, errors='coerce')

    if pd.isna(start_date) or pd.isna(end_date):
        st.error("Invalid date format provided.")
        # return

                # Load the Excel file
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)

                    # Process the 'Current week' sheet
        ws = wb['Current week']

                    # Generate the dates and write them to cells A1 to A7
        for i, date in enumerate(generate_dates(start_date, end_date)):
            cell_a = f'A{i+1}'
            cell_b = f'B{i+1}'
            ws[cell_a] = date
            ws[cell_b] = "Yes"

                    # Save changes
        wb.save(file_path)

        # st.write(f"Dates from {start_date_str} to {end_date_str} have been written to cells A1 to A7, and 'Yes' to column B.")

                    # Process the 'Raw Data' sheet
        ws = wb['Raw Data']

        df2_path = r'\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet\uploaded_file2.xlsx'
        df2 = pd.read_excel(df2_path)

                    # Convert date column in df2 to datetime, handling errors
        df2['Date'] = pd.to_datetime(df2['Date'], errors='coerce')

                    # Font settings
        font = Font(name='Trebuchet MS', size=8)

                    # Copy data from df2 to columns D and G in the Excel sheet
        first_blank_d = find_first_empty_row(ws, 'D')
        first_blank_g = find_first_empty_row(ws, 'G')

        for i, (date, agent) in enumerate(zip(df2['Date'], df2['Agent']), start=first_blank_d):
            if pd.notna(date):  # Ensure date is valid
                cell_d = ws[f'D{i}']
                cell_d.value = date
                cell_d.font = font

            cell_g = ws[f'G{i}']
            cell_g.value = agent
            cell_g.font = font

                    # List of DataFrame columns to copy and corresponding Excel columns
        columns_to_copy = [
                        ('SkillGroupName', 'E'),
                        ('CallsHandled', 'H'),
                        ('OutExtnCalls', 'I'),
                        ('InternalCalls', 'J'),
                        ('RedirectCalls', 'K'),
                        ('AHT', 'L'),
                        ('AnswerWaitTime', 'M'),
                        ('TalkTime', 'N'),
                        ('HoldTime', 'O'),
                        ('ReservedTime', 'P'),
                        ('AgentBusyOtherTime', 'Q'),
                        ('WorkNotReadyTime', 'R'),
                        ('AgentAvailTime', 'S'),
                        ('AgentLoggedOnTime', 'T'),
                        ('Assists', 'U'),
                        ('TransferOutCalls', 'V'),
                        ('ConferenceOutCalls', 'W'),
                        ('ConsultativeCalls', 'X'),
                        ('InCallsOnHold', 'Y')
                    ]

                    # Copy each column from df2 to the corresponding column in Excel
        for df_column, excel_column in columns_to_copy:
            copy_data_to_excel(df2, ws, df_column, excel_column, font)

                    # Define the formulas
        formulas = {
                        'F': '=VLOOKUP(G{},Agent!A:B,2,FALSE)',
                        'A': '=IFERROR(VLOOKUP(D{},\'Current week\'!$A:$B,2,FALSE),"No")',
                        'C': '=TEXT(D{},"MMMMMMMMMMM")',
                        'B': '=VLOOKUP(C{},\'Current week\'!$K$2:$L$13,2,FALSE)',
                        'AA': '=Z{}*Y{}',
                        'AB': '=IFERROR(VLOOKUP(E{}&F{},Teams!$A:$D,4,FALSE),"NA")'
                    }

                    # Define the formatting
        formatting = {
                        'F': {'font': Font(name='Trebuchet MS', size=8), 'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')},
                        'A': {'font': Font(name='Trebuchet MS', size=8), 'fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')},
                        'C': {'font': Font(name='Trebuchet MS', size=8), 'fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')},
                        'B': {'font': Font(name='Trebuchet MS', size=8), 'fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')},
                        'AA': {'font': Font(name='Trebuchet MS', size=8), 'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')},
                        'AB': {'font': Font(name='Trebuchet MS', size=8), 'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')}
                    }

                    # Apply the formulas and formatting
        stop_row_e = len(ws['E'])
        for i, cell in enumerate(ws['E'], start=1):
            if cell.value is None or cell.value == "":
                stop_row_e = i - 1
                break

        for col, start_row in zip(['F', 'A', 'C', 'B', 'AA', 'AB'], [find_first_empty_row(ws, col) for col in ['F', 'A', 'C', 'B', 'AA', 'AB']]):
            if start_row is not None:
                for row in range(start_row, stop_row_e + 1):
                    cell = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col))
                    if col == 'AA':
                        cell.value = formulas[col].format(row, row)
                    elif col == 'AB':
                        cell.value = formulas[col].format(row, row)
                    else:
                        cell.value = formulas[col].format(row)
                    cell.font = formatting[col]['font']
                    cell.fill = formatting[col]['fill']

                    # Save the updated workbook
        wb.save(file_path)

        wb = openpyxl.load_workbook(file_path)
        ws = wb['Abandon Calls data']
        df1_path = r'\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet\uploaded_file1.xlsx'
        df1 = pd.read_excel(df1_path)
                # Font settings
        font = Font(name='Trebuchet MS', size=8)

                # Dictionary to map columns in the Excel sheet to DataFrame columns
        column_mapping = {
                    'D': 'SkillName',
                    'E': 'Date',
                    'H': 'Handled',
                    'K': 'Abandon',
                    'S': 'RONA'
                }

                # Copy data from df1 to specified columns in the Excel sheet
        for excel_col, df_col in column_mapping.items():
            first_empty_row = find_first_empty_row(ws, excel_col)

            for i, value in enumerate(df1[df_col], start=first_empty_row):
                cell = ws[f'{excel_col}{i}']

                if df_col == 'Date':  # Convert date to datetime format
                    cell.value = pd.to_datetime(value)
                else:
                    cell.value = value

                cell.font = font

                # Save the updated workbook
        wb.save(file_path)
        # st.write(f"Data from df1 has been pasted into the specified columns with font Trebuchet MS and size 8.")

                # Process the 'Abandon Calls data' sheet further
        ws = wb['Abandon Calls data']

                # Font and fill settings
        font = Font(name='Trebuchet MS', size=8)
        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

                # Formulas and their respective columns
        formulas = {
                    'A': '=IFERROR(VLOOKUP(E{},\'Current week\'!$A:$B,2,FALSE),"No")',
                    'B': '=VLOOKUP(C{},\'Current week\'!$K$2:$L$13,2,FALSE)',
                    'C': '=TEXT(E{},"MMMMMMMMMMM")',
                    'W': '=VLOOKUP(D{},Teams!B:D,3,FALSE)'
                }

                # Start rows for each column
        start_rows = {col: find_first_empty_row(ws, col) for col in ['A', 'B', 'C', 'W']}

                # Stop row for column E
        stop_row_e = find_stop_row(ws, 'D')

                # Apply the formulas from the start row to the stop row
        for col, formula in formulas.items():
            start_row = start_rows[col]
            for row in range(start_row, stop_row_e + 1):
                cell = ws.cell(row=row, column=openpyxl.utils.column_index_from_string(col))  # Convert col to index
                cell.value = formula.format(row)
                cell.font = font
                cell.fill = fill

                # Save the updated workbook
        wb.save(file_path)

        st.success('Call Volume Data - Base Data - working file successfully updated.')

st.page_link("streamlit_app.py", label="Home Page",icon="🏠")
st.page_link("pages/page1.py", label="Abandon Call & Combined Call volume Excel file process", icon="1️⃣")
st.page_link("pages/page2.py", label="Call volume - Base Data - working", icon="2️⃣")
st.page_link("pages/page3.py",label="Veronica & Curtis File - working", icon="3️⃣")
