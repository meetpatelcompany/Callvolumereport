import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
import os
import openpyxl
from datetime import datetime

import pyautogui

def process_excel_file(file_path, sheet_name, date_column='Date'):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Delete the first row
    ws.delete_rows(1)
    
    # Identify the column index for the 'Date' column
    date_col_index = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == date_column:
            date_col_index = col_idx
            break

    if date_col_index is None:
        raise ValueError(f"Column '{date_column}' not found in the sheet.")

    # List to keep track of rows to delete
    rows_to_delete = set()
    
    # Iterate through rows to find empty cells in the 'Date' column
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=date_col_index, max_col=date_col_index, values_only=True), start=2):
        if row[0] is None:
            rows_to_delete.update(range(row_idx, min(row_idx + 6, ws.max_row + 1)))  # Add 5 more rows to the deletion set
    
    # Delete rows in reverse order to avoid affecting the indices of rows yet to be deleted
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    # Save the changes to the Excel file
    wb.save(file_path)

def filter_dataframe_with_openpyxl(file_path, sheet_name, conditions, exclude_value=None, exclude_column=None):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    
    # Exclude specific value if provided
    if exclude_value is not None and exclude_column is not None:
        df = df[df[exclude_column] != exclude_value]
    
    filtered_df = df[conditions]
    
    # Clear the worksheet and write the filtered data back
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(file_path)

# Function to generate dates between start and end date
def generate_dates(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)

# Function to find the first empty row in a column
def find_first_empty_row(ws, column):
    for row in range(1, ws.max_row + 1):
        if ws[f'{column}{row}'].value is None:
            return row
    return ws.max_row + 1

# Function to copy data from DataFrame to Excel
def copy_data_to_excel(df, ws, df_column, excel_column, font):
    first_empty_row = find_first_empty_row(ws, excel_column)
    for i, value in enumerate(df[df_column], start=first_empty_row):
        cell = ws[f'{excel_column}{i}']
        cell.value = value
        cell.font = font

def find_stop_row(sheet, col):
    for i, cell in enumerate(sheet[col], start=1):
        if cell.value is None or cell.value == "":
            return i - 1
    return len(sheet[col])

def get_quarter(month):
    if month in [1, 2, 3]:
        return 'Q1'
    elif month in [4, 5, 6]:
        return 'Q2'
    elif month in [7, 8, 9]:
        return 'Q3'
    else:
        return 'Q4'

# Function to get month name from month number
def get_month_name(month):
    return datetime(2024, month, 1).strftime('%B')

st.subheader("Instructions for Page 3",divider=True)

st.markdown("This process devide into 4 parts")
st.markdown("1. Choose Start and end Date for Previous week")
st.markdown("2. Press Execute button and wait till process complete")
st.markdown("3. Open Base data - working file")
st.markdown("4. Open Curtis and Veronica file it will ask for update link so press that and choose Base data - working file and copy paste as value and save it")
# st.markdown("3. Press Execute button and wait till process complete")
# st.markdown("4. You can find modified files in :blue-background[\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet]")

start_date_str = st.date_input('Start Date', datetime(2024, 7, 15))
end_date_str = st.date_input('End Date', datetime(2024, 7, 21))
start_date = pd.to_datetime(start_date_str, errors='coerce')
end_date = pd.to_datetime(end_date_str, errors='coerce')
month = end_date.month
quarter = get_quarter(month)
month_name = get_month_name(month)

# Start date for Sheet 2


# Open the Excel file

if st.button('Execute'):
    file_path_old = r"\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet\Call Volume Data - Base Data - Working - CUIC.xlsx"
    wb_old = openpyxl.load_workbook(file_path_old)
    file_path = r'\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet\Call Volume Data - Weekly Report - Curtis – SCC – Sales Contact Centre - Working - CUIC.xlsx'
    wb = openpyxl.load_workbook(file_path)

# Sheet 1: Summary by Gates
    sheet1 = wb['Summary by Gates']
    sheet1['B1'] = quarter
    sheet1['B2'] = month_name

# Sheet 2: Summary-Week
    sheet2 = wb['Summary-Week']
    date_range_str = f'{start_date.strftime("%d")}-{end_date.strftime("%d")} {month_name}'
    sheet2['B1'] = date_range_str
# sheet2['B1'] = f'{start_date_str} - {end_date_str} {month_name}'

# Sheet 3: Summary-MTD
    sheet3 = wb['Summary-MTD']
    for row in range(2, 43):
        sheet3[f'C{row}'] = month_name

# Sheet 4: Summary-QTD
    sheet4 = wb['Summary-QTD']
    for row in range(2, 43):
        sheet4[f'C{row}'] = quarter

# Save the modified Excel file
    wb.save(file_path)

    file_path = r'\\ETONFS1\Departments\Sales Analytics\AdHoc\Meet\Call Volume Data - Weekly Report - Veronica – CS – Customer Solutions  - Working - CUIC.xlsx'
    wb = openpyxl.load_workbook(file_path)

# Sheet 1: Summary by Gates
    sheet1 = wb['Summary by Gates']
    sheet1['B1'] = quarter
    sheet1['B2'] = month_name

# Sheet 2: Summary-Week
    sheet2 = wb['Summary-Week']
    date_range_str = f'{start_date.strftime("%d")}-{end_date.strftime("%d")} {month_name}'
    sheet2['B1'] = date_range_str
# sheet2['B1'] = f'{start_date_str} - {end_date_str} {month_name}'

# Sheet 3: Summary-MTD
    sheet3 = wb['Summary-MTD']
    for row in range(2,10):
        sheet3[f'C{row}'] = month_name

# Sheet 4: Summary-QTD
    sheet4 = wb['Summary-QTD']
    for row in range(2, 10):
        sheet4[f'C{row}'] = quarter

# Save the modified Excel file
    wb.save(file_path)
    st.success('Call Volume Report for Curtis & Veronica has been Updated.')

st.page_link("main.py", label="Home Page",icon="🏠")
st.page_link("pages/page1.py", label="Abandon Call & Combined Call volume Excel file process", icon="1️⃣")
st.page_link("pages/page2.py", label="Call volume - Base Data - working", icon="2️⃣")
st.page_link("pages/page3.py",label="Veronica & Curtis File - working", icon="3️⃣")


