import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
import os

import pyautogui

def process_excel_file(file_path, sheet_name, date_column='Date'):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Delete the first row
    ws.delete_rows(1)
    
    # Identify the column index for the 'Date' column
    date_col_index = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == date_column:
            date_col_index = col_idx
            break

    if date_col_index is None:
        raise ValueError(f"Column '{date_column}' not found in the sheet.")

    # List to keep track of rows to delete
    rows_to_delete = set()
    
    # Iterate through rows to find empty cells in the 'Date' column
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=date_col_index, max_col=date_col_index, values_only=True), start=2):
        if row[0] is None:
            rows_to_delete.update(range(row_idx, min(row_idx + 6, ws.max_row + 1)))  # Add 5 more rows to the deletion set
    
    # Delete rows in reverse order to avoid affecting the indices of rows yet to be deleted
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    
    # Save the changes to the Excel file
    wb.save(file_path)

def filter_dataframe_with_openpyxl(file_path, sheet_name, conditions, exclude_value=None, exclude_column=None):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    
    # Exclude specific value if provided
    if exclude_value is not None and exclude_column is not None:
        df = df[df[exclude_column] != exclude_value]
    
    filtered_df = df[conditions]
    
    # Clear the worksheet and write the filtered data back
    ws.delete_rows(1, ws.max_row)
    for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(file_path)

# Function to generate dates between start and end date
def generate_dates(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)

# Function to find the first empty row in a column
def find_first_empty_row(ws, column):
    for row in range(1, ws.max_row + 1):
        if ws[f'{column}{row}'].value is None:
            return row
    return ws.max_row + 1

# Function to copy data from DataFrame to Excel
def copy_data_to_excel(df, ws, df_column, excel_column, font):
    first_empty_row = find_first_empty_row(ws, excel_column)
    for i, value in enumerate(df[df_column], start=first_empty_row):
        cell = ws[f'{excel_column}{i}']
        cell.value = value
        cell.font = font

def find_stop_row(sheet, col):
    for i, cell in enumerate(sheet[col], start=1):
        if cell.value is None or cell.value == "":
            return i - 1
    return len(sheet[col])




st.title("Call Volume Report - Weekly")
st.subheader("Instructions for New process of Call Volume Report",divider=True)

st.markdown("This process devide into 4 parts")
st.markdown("1. Page 1 - Process of Abandon Calls and Combined Calls filering.")
st.markdown("2. Page 2 - Updating Base Data - working file with Previous week Data")
st.markdown("3. Page 3 - Updating the Veronica & Curtis file")
st.markdown("4. Once process is finish Press reset button to reset everything")
if st.button("Reset"):
    pyautogui.hotkey("ctrl","F5")

# st.divider()
# st.markdown(":blue-background[1. Filtering process for Abandon Calls and Call Volumes Combined]")

            
st.page_link("main.py", label="Home Page",icon="🏠")
st.page_link("pages/page1.py", label="Abandon Call & Combined Call volume Excel file process", icon="1️⃣")
st.page_link("pages/page2.py", label="Call volume - Base Data - working", icon="2️⃣")
st.page_link("pages/page3.py",label="Veronica & Curtis File - working", icon="3️⃣")
